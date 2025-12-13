# OpenPyXL Example 
# pip install openpyxl  
import openpyxl  
import os  
os.chdir('c:\\users\al\\documents')  

workbook = openpyxl.load_workbook('example.xlsx') # relative filepath so going to look for it in the current working directory  
workbook.get_sheet_names()  
sheet = workbook.get_sheet_by_name('Sheet1')  
cell = sheet ['A1']	# evaluates to a cell object  
cell.value # to get the actual value in that cell  
str(sheet['A1']) # get string value of cell  
sheet.cell(row = 1, column = 2)	# evaluates to same as above but may be preferred in some functions  

# Editing Excel Spreadsheets  
import openpyxl  
wb = openpyxl.Workbook()  
wb.get_sheet_names()  
sheet = wb.get_sheet_by_name('Sheet')  

sheet['A1'] = 42  
sheet['A2'] = 'Hello' # saved in computer's memory  

import os   
os.chdir('c:\\Users\\Al\\Documents')  
wb.save() # to save to hardrive  
openpyx.load_workbook('abc.xslx')  

# To create a sheet  
sheet2= wb.create_sheet() # creates sheet to end of workbook by default  
wb.get_sheet_names()  
sheet2.title = 'My New Sheet Name'  
wb.get_sheet_names()  
wb.save('example2.xlsx')  
wb.create_sheet(index  = 0, title  = 'My Other Sheet') # creates sheet at beginning and changes name  